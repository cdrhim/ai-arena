from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


CATEGORY_RULES = [
    ("Healthcare / Bio", ["의료", "헬스", "병원", "의학", "제약", "암", "항암", "바이오", "건강", "약국", "디지털헬스", "환자", "임상", "통증", "항염증", "복약"]),
    ("Legal / IP", ["법률", "변호사", "변리사", "청구항", "OA 대응", "IP/Legal", "Legal Tech", "인벤팁스"]),
    ("Finance / Investment", ["금융", "퀀트", "펀드", "대출", "은행", "자산", "신탁", "캐피탈", "보험", "Fintech", "재무", "리서치", "애널리스트", "온체인"]),
    ("Real Estate / PropTech", ["부동산", "상업용", "담보", "공인중개사", "PropDeal", "미쓰이부동산"]),
    ("Creative / Art", ["미술", "아트", "Art Basel", "갤러리", "작품", "작가", "수집가"]),
    ("Commerce / Retail", ["커머스", "이커머스", "리테일", "유통", "SKU", "재고", "쇼핑", "브랜드", "주문", "판매사", "쿠팡", "스토어"]),
    ("Marketing / AdTech", ["광고", "마케팅", "인플루언서", "캠페인", "ROAS", "SNS", "콘텐츠 제작", "광고주", "브랜디드"]),
    ("Media / Entertainment", ["엔터", "영상", "게임", "음원", "캐스팅", "VFX", "캐릭터", "크리에이터", "버추얼", "음악", "영화", "퍼블리셔", "세계관", "VR 콘텐츠"]),
    ("Education / Research", ["교육", "학습", "학생", "대학", "학교", "입시", "학종", "연구", "논문", "학술", "멘토", "유아", "어린이", "교사", "시간표", "SAT"]),
    ("HR / Workforce", ["채용", "HR", "인재", "근태", "근로계약", "급여", "인력관리", "출근부", "구인구직", "파트타임"]),
    ("Travel / Hospitality", ["여행", "호텔", "게스트하우스", "숙박", "OTA", "관광", "예약", "호스트"]),
    ("Food / F&B", ["외식", "식당", "프랜차이즈", "카페", "커피", "매장", "POS", "미쉐린", "F&B", "식품"]),
    ("Robotics / Mobility", ["로봇", "드론", "임베디드", "CAN", "자율주행", "차량", "모빌리티", "전장", "항공", "에어웨이", "PAV", "사족보행", "틸트로터"]),
    ("Manufacturing / Materials", ["소재", "재활용", "플라스틱", "제조", "공장", "공급사", "RFQ", "물성", "FT-IR", "TGA", "DSC", "양산", "하드웨어", "펌웨어", "시제품"]),
    ("Logistics / Supply Chain", ["물류", "창고", "Fulfillment", "배송", "공급망", "LPR", "주차장"]),
    ("Security / Compliance", ["보안", "OWASP", "취약점", "위험도", "개인정보", "프라이버시", "인증", "CASA"]),
    ("Beauty / Wellness", ["뷰티", "피부", "미용", "웰니스", "건강관리기기"]),
    ("Climate / Energy", ["탄소", "에너지", "환경", "기후", "재생", "폐기물"]),
    ("Developer / AI Infrastructure", ["LLM", "에이전트", "RAG", "벡터", "Qdrant", "eval", "API", "개발자", "노코드", "프롬프트", "모델", "KV", "LongBench", "어텐션", "MCP", "오케스트레이션", "파인튜닝"]),
    ("Operations / Productivity", ["SaaS", "자동화", "업무", "대시보드", "워크플로", "생산성", "문서", "리포트", "운영", "PM", "PRD"]),
]

FUNCTION_RULES = [
    ("AI Agents", ["에이전트", "Agent", "오케스트레이션", "MCP"]),
    ("Computer Vision", ["비전", "CV", "이미지", "영상", "차트 파싱", "물성 분석", "LPR", "VFX"]),
    ("Forecasting", ["예측", "forecast", "demand", "수요", "Croston"]),
    ("RAG", ["RAG", "검색", "증거검색", "인용검증"]),
    ("LLM", ["LLM", "GPT", "Claude", "프롬프트", "파인튜닝", "모델"]),
    ("Workflow Automation", ["자동화", "워크플로", "파이프라인", "대행", "리포트"]),
    ("Data Analytics", ["데이터", "분석", "analytics", "대시보드"]),
    ("Marketplace", ["거래", "매칭", "플랫폼", "마켓", "커머스"]),
    ("Robotics / Hardware", ["로봇", "하드웨어", "펌웨어", "임베디드", "센서", "기기"]),
    ("Security", ["보안", "취약점", "위험도", "인증"]),
    ("Recommendation", ["추천", "매칭", "개인화"]),
    ("Document AI", ["문서", "PDF", "계약", "특허", "논문", "리포트"]),
]


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\r", " ").replace("\n", " / ")).strip()


def normalize_label(label: str) -> str:
    label = clean(label)
    label = re.sub(r"\(ROW#\d+\)", "", label)
    label = re.sub(r"\([^)]*\)$", "", label)
    return re.sub(r"\s+", "", label).lower()


def strip_company_prefix(company: str) -> str:
    company = clean(company)
    for prefix in ["주식회사 ", "주식회사", "(주)", "주)"]:
        if company.startswith(prefix):
            company = company[len(prefix) :].strip()
    return company


def split_company_product(label: str) -> tuple[str, str, str, str]:
    raw = clean(label)
    row_match = re.search(r"ROW#(\d+)", raw)
    source_row = row_match.group(1) if row_match else ""
    founder = ""
    founder_match = re.search(r"\(([^)]*)\)\s*$", raw)
    if founder_match and "ROW#" not in founder_match.group(1):
        founder = clean(founder_match.group(1))
    label_no_paren = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
    parts = [clean(part) for part in label_no_paren.split("/", 1)]
    company = strip_company_prefix(parts[0]) if parts else label_no_paren
    product = parts[1] if len(parts) > 1 and parts[1] else company
    return company or label_no_paren, product, founder, source_row


def public_sites_by_company(workbook) -> dict[str, str]:
    site_by_key = {}
    for sheet in ["0617", "0609"]:
        if sheet not in workbook.sheetnames:
            continue
        rows = list(workbook[sheet].iter_rows(values_only=True))
        for row in rows[1:]:
            vals = [clean(value) for value in row]
            if not vals or not vals[0]:
                continue
            site = vals[5] if sheet == "0617" and len(vals) > 5 else ""
            if not site and len(vals) > 4:
                site = vals[4]
            if "batch-zeta" in site or "/admin/applications/" in site:
                site = ""
            if site and "." in site:
                if not re.match(r"https?://", site, re.I):
                    site = f"https://{site}"
                site_by_key[normalize_label(vals[0])] = site
    return site_by_key


def infer_category(text: str) -> str:
    lower = text.lower()
    for category, keywords in CATEGORY_RULES:
        if any(keyword.lower() in lower for keyword in keywords):
            return category
    return "AI / SaaS"


def infer_functions(text: str, category: str) -> list[str]:
    lower = text.lower()
    functions = []
    for label, keywords in FUNCTION_RULES:
        if any(keyword.lower() in lower for keyword in keywords):
            functions.append(label)
    if not functions:
        functions.append("Workflow Automation" if "AI" in category or "SaaS" in category else "AI Enablement")
    if "Data Analytics" not in functions and any(keyword in lower for keyword in ["데이터", "지표", "매출", "mrr", "gmv", "사용자"]):
        functions.append("Data Analytics")
    return functions[:5]


def infer_stage(traction: str) -> str:
    lower = traction.lower()
    if any(keyword in lower for keyword in ["연구단계", "유저 0", "매출 0", "closed beta 전", "출시 전"]):
        return "Research"
    if any(keyword in lower for keyword in ["누적 거래액 40억", "630억", "매출 20억", "매출 7.46억", "mau 약 80만", "가입 9,034", "mrr 3,500만", "5.2억", "4.7억"]):
        return "Growth"
    if any(keyword in lower for keyword in ["매출", "mrr", "유료", "계약", "고객", "가입", "mau", "사용자", "다운로드", "gmv", "거래액", "paid", "loi"]):
        return "Seed"
    if any(keyword in lower for keyword in ["mvp", "베타", "poc", "파일럿", "실증", "데모", "사전예약"]):
        return "Pre-Seed"
    return "Pre-Seed"


def traction_score(traction: str, strength: str) -> int:
    lower = traction.lower()
    score = 20
    for keyword, points in [
        ("mrr", 12),
        ("매출", 12),
        ("유료", 10),
        ("계약", 10),
        ("고객", 7),
        ("mau", 7),
        ("가입", 6),
        ("사용자", 5),
        ("다운로드", 6),
        ("gmv", 9),
        ("거래액", 9),
        ("tips", 4),
        ("특허", 4),
        ("loi", 5),
        ("poc", 4),
        ("파일럿", 4),
    ]:
        if keyword in lower:
            score += points
    if "traction" in strength.lower():
        score += 8
    if "technical" in strength.lower():
        score += 3
    return max(18, min(96, score))


def tagline_for(product: str, category: str) -> str:
    base = {
        "Healthcare / Bio": "AI-enabled healthcare and bio workflow validation.",
        "Legal / IP": "AI workflow for legal/IP review and execution.",
        "Finance / Investment": "AI-assisted finance, investment, or risk workflow.",
        "Real Estate / PropTech": "AI workflow for real estate, property finance, and asset operations.",
        "Creative / Art": "AI product for creative discovery, art, and cultural workflows.",
        "Commerce / Retail": "AI product for commerce, retail, and revenue operations.",
        "Marketing / AdTech": "AI product for marketing execution and growth workflows.",
        "Media / Entertainment": "AI product for content, media, and entertainment workflows.",
        "Education / Research": "AI product for learning, research, and expert workflows.",
        "HR / Workforce": "AI workflow for workforce, HR, and operations teams.",
        "Travel / Hospitality": "AI product for travel, hospitality, and guest operations.",
        "Food / F&B": "AI workflow for F&B operators and store teams.",
        "Robotics / Mobility": "AI-enabled robotics, mobility, or hardware workflow.",
        "Manufacturing / Materials": "AI product for industrial, manufacturing, and materials workflows.",
        "Logistics / Supply Chain": "AI workflow for logistics, fulfillment, and inventory operations.",
        "Security / Compliance": "AI product for security, risk, and compliance workflows.",
        "Beauty / Wellness": "AI-enabled wellness, beauty, or health product workflow.",
        "Climate / Energy": "AI workflow for climate, energy, and sustainability operations.",
        "Developer / AI Infrastructure": "AI infrastructure product for builders and operators.",
        "Operations / Productivity": "AI-native workflow product for business operations.",
    }.get(category, "AI-native product from the Spark Claw applicant pool.")
    return f"{product}: {base}"[:160]


def slug_id(index: int) -> str:
    return f"sparkclaw-applicant-{index:03d}"


def build_seed(input_path: Path) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(input_path, data_only=True, read_only=False)
    site_by_key = public_sites_by_company(workbook)
    rows = []
    for row in workbook["0626"].iter_rows(values_only=True):
        vals = [clean(value) for value in row]
        if any(vals):
            rows.append(vals)

    startups = []
    category_counts = {}
    for index, vals in enumerate(rows[1:], start=1):
        name_field = vals[0]
        strength = vals[1] if len(vals) > 1 else ""
        domain = vals[2] if len(vals) > 2 else ""
        traction = vals[3] if len(vals) > 3 else ""
        company, product, founder, source_row = split_company_product(name_field)
        text = " ".join([name_field, strength, domain, traction])
        category = infer_category(text)
        functions = infer_functions(text, category)
        stage = infer_stage(traction)
        score = traction_score(traction, strength)
        startup_id = slug_id(index)
        tags = []
        for token in re.split(r"[·,;/\s]+", strength):
            token = clean(token)
            if token and token not in tags:
                tags.append(token)
        for function in functions:
            if function not in tags:
                tags.append(function)
        category_counts[category] = category_counts.get(category, 0) + 1
        startups.append(
            {
                "id": startup_id,
                "name": company,
                "founder": founder or "Applicant team",
                "category": category,
                "functions": functions,
                "stage": stage,
                "region": "Korea",
                "affiliation": "SparkClaw Applicant",
                "tagline": tagline_for(product, category),
                "description": f"{domain[:377]}..." if len(domain) > 380 else domain,
                "traction": f"{traction[:417]}..." if len(traction) > 420 else traction,
                "upvotes": int(score + index % 17),
                "demoRequests": int(max(2, min(35, score // 4 + index % 5))),
                "corporateInterest": int(max(1, min(20, score // 6 + index % 4))),
                "investorInterest": int(max(1, min(20, score // 7 + (index + 2) % 4))),
                "benchmarkScore": round(max(55, min(96, score + 8 + (index % 9) * 0.7)), 1),
                "tags": tags[:10],
                "sourceSheet": "0626",
                "sourceRow": source_row or str(index + 1),
                "products": [
                    {
                        "id": f"{startup_id}-product",
                        "name": product,
                        "type": category,
                        "url": site_by_key.get(normalize_label(name_field), ""),
                        "demoVideoUrl": "",
                        "useCases": functions[:3],
                        "upvotes": int(max(1, score - 5 + index % 13)),
                        "reviews": int(max(1, min(24, score // 8))),
                    }
                ],
            }
        )
    return startups, sorted(category_counts)


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: generate_sparkclaw_applicant_seed.py <workbook.xlsx> <output.mjs>")
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    startups, business_areas = build_seed(input_path)
    output_path.write_text(
        "// Generated from Spark Claw Program 지원자 현황 시트 (벤처본부 공유).xlsx, sheet 0626.\n"
        "// Do not expose internal batch-zeta application URLs in this seed data.\n"
        f"export const SPARKCLAW_APPLICANT_STARTUPS = {json.dumps(startups, ensure_ascii=False, indent=2)};\n\n"
        f"export const SPARKCLAW_BUSINESS_AREAS = {json.dumps(business_areas, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )
    counts = {}
    for startup in startups:
        counts[startup["category"]] = counts.get(startup["category"], 0) + 1
    print(f"wrote {output_path} startups={len(startups)} business_areas={len(business_areas)}")
    for category, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        print(f"{category}: {count}")


if __name__ == "__main__":
    main()
